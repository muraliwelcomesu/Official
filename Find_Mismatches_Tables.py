'''
Created on Jun 6, 2019

@author: MURENGAR
This pgm will compare the tables in two schemas and identify the mismatches in columns(with datatypes).
Output will be stored in an excel.
'''
import cx_Oracle
import os,sys
import openpyxl

''' This function will return a dict with the column details of all Tables passed from the schema passed as input'''
def get_table_details_schema(p_conn_str,p_list_tables,l_include_data_type):
    connection1 = cx_Oracle.connect(p_conn_str)
    cur1 = connection1.cursor()
    l_dict1 = {}
    #l_list_tables = ['ERTB_MSGS']
    for i in p_list_tables:
        result1 = []
        if l_include_data_type == 'Y':
            l_str = 'SELECT table_name||'+ "'.'"+ '||column_name||' + "'::'" + '||data_type||' + "'('" + '||char_length||'+ "')'" + ' from user_tab_columns where table_name = '+ "'" + i  +"'" + 'order by column_id'
        else:
            l_str = 'SELECT table_name||'+ "'.'"+ '||column_name '  + ' from user_tab_columns where table_name = '+ "'" + i  +"'" + 'order by column_id'
        #print(l_str)
        rs1 = cur1.execute(l_str)
        for res1 in cur1:
            result1.append(res1[0])
        l_dict1[i] = result1
    cur1.close()
    connection1.close()
    return l_dict1
''' this function compares two dictionaries and returns list of key values from first dict which is missing in second dict 
and also common keys from first dict'''
def fn_classify_java_tbls(l_dict_Java,l_dict_Plsql):
    l_dict_java_2 = {}
    l_lst_newIC_tbls = []
    for key,values in l_dict_Java.items():
        if len(l_dict_Plsql[key]) < 1:
            l_lst_newIC_tbls.append(key)
        else:
            l_dict_java_2[key] = l_dict_Java[key]
    return l_lst_newIC_tbls,l_dict_java_2

''' Given a list of specific format , this function would convert it  to a dict '''
def split_list_dict(p_list):
    l_dict1 = {}
    for i in p_list:
        l_dict1[i.split('::')[0]] = i.split('::')[1]
    return l_dict1
''' This function compares two dicts and returns the difference in values for same key in both dict in format key::dict1::dict2
and also key missing in dict1 and those missing in dict2'''
def compare_dicts(p_dict1,p_dict2):
    l_diff_list = []
    l_missing_list_2 = []
    l_missing_list_1 = []
    for key,value in p_dict1.items():
        if key not in p_dict2:
            l_missing_list_2.append(key)
        elif p_dict1[key] != p_dict2[key]:
            l_str = key + '::' + p_dict1[key] + '::' + p_dict2[key]
            l_diff_list.append(l_str)
    for key,value in p_dict2.items():
        if key not in p_dict1:
            l_missing_list_1.append(key)
    return l_diff_list,l_missing_list_2,l_missing_list_1
''' This function executes a plsql query and return result as list '''
def execute_query(p_conn_Str,p_query_str):
    connection1 = cx_Oracle.connect(p_conn_Str)
    cur1 = connection1.cursor()
    result1 = []
    rs1 = cur1.execute(p_query_str)
    for res1 in cur1:
        result1.append(res1[0])
    cur1.close()
    connection1.close()
    return result1
''' This function compares two dict and returns dict of differences in them
for each key in first dict, it returns a list of 2 elements 
list[0] - mismatch of columns in dict1 and dict2 for a key
list[1] - values missing in dict2
list[2] values missing in dict1 '''

def find_diff_Lists(p_dict1,p_dict2):
    print('inside find_diff_Lists')
    l_final_dict = {}
    l_final_list = []
    for key,value in p_dict1.items():
        #print('key ',key)
        dict1 = {}
        dict2 = {}
        list1 = []
        list2 = []
        list_tmp_1 = []
        l_diff_list = []
        l_missing_list_2 = []
        l_missing_list_1 = []
        if key in p_dict2:
            list1 = (sorted(p_dict1[key])) 
            list2 = (sorted(p_dict2[key]))
            dict1 = split_list_dict(list1)
            dict2 = split_list_dict(list2)
            l_diff_list,l_missing_list_2,l_missing_list_1 = compare_dicts(dict1,dict2)
            l_final_list.append(l_diff_list)
            l_final_list.append(l_missing_list_2)
            l_final_list.append(l_missing_list_1)
        l_final_dict[key] = l_final_list
        l_final_list = []
    return l_final_dict
''' Write the output of dict to an excel '''
def write_to_excel(l_new_tables_java,l_final_dict):
    wb1 = openpyxl.Workbook()
    sheet = wb1.create_sheet()
    sheet.title = 'Mismatches'
    sheet['A1'] = 'TableName'
    sheet['B1'] = 'MISMATCH_COLNAME:PLSQL_SCHEMA::ICJAVA_SCHEMA'
    sheet['C1'] = 'MISSING_COLS_PLSQL'
    sheet['D1'] = 'MISSING_COLS_JAVA'
    l_cols = ['B','C','D']
    l_colidx = 0
    l_rownum = 1
    for key,values in l_final_dict.items():
        if len(values[0]) > 0:
            for i in values[0]:
                l_rownum += 1
                sheet['A' + str(l_rownum)].value = key
                sheet['B' + str(l_rownum)].value = i
                
        if len(values[1]) > 0:
            for i in values[1]:
                l_rownum += 1
                sheet['A' + str(l_rownum)].value = key
                sheet['C' + str(l_rownum)].value = i
                
        if len(values[2]) > 0:
            for i in values[2]:
                l_rownum += 1
                sheet['A' + str(l_rownum)].value = key
                sheet['D' + str(l_rownum)].value = i
                
    sheet1 = wb1.create_sheet()
    sheet1.title = 'New Tables'
    sheet1['A1'] = 'TableName'
    l_rownum = 1
    for i in l_new_tables_java:
        l_rownum += 1
        sheet1['A' + str(l_rownum)].value = i
    wb1.remove(wb1['Sheet'])
    wb1.save('C:\\Users\\MURENGAR\\Desktop\\Output\\Mismatches.xlsx')  
    
if __name__ == "__main__":
    print('starting...')
    p_conn_str_14_2 = Config.Connstr1
    p_conn_str_Java = Config.Connstr2
    l_query_Str = 'select  table_name From user_tables order by table_name'
    l_include_data_type = 'Y' 
    l_final_dict = {}
    l_list_tabls  = execute_query(p_conn_str_Java,l_query_Str)
    l_dict_Java_tmp = get_table_details_schema(p_conn_str_Java,l_list_tabls,l_include_data_type)
    l_dict_Plsql = get_table_details_schema(p_conn_str_14_2,l_list_tabls,l_include_data_type)
    l_new_tables_java,l_dict_Java = fn_classify_java_tbls(l_dict_Java_tmp,l_dict_Plsql)
    l_final_dict = find_diff_Lists(l_dict_Java,l_dict_Plsql )
    write_to_excel(l_new_tables_java,l_final_dict)
    print('Completed')
