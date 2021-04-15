import openpyxl
import os
import pandas as pd

def convert_excel_txt(path):
    os.chdir(path)
    lst_excel = sorted([x for x in  os.listdir() if x.endswith('.xlsx')])
    for ExcelName in lst_excel:
        wb = openpyxl.load_workbook(ExcelName)
        lst_sheets = wb.sheetnames
        for sheetname  in lst_sheets:
            df  = pd.read_excel(ExcelName,sheet_name = sheetname)
            filename = '{}_{}.txt'.format(os.path.splitext(ExcelName)[0],sheetname)
            with open(filename, 'w') as f:
                f.write(df.to_string(header = True, index = False))
            print('File {} created successfully'.format(filename))
    print('Completed')
    
def convert_excel_txt2(path):
    os.chdir(path)
    dest_folder = os.path.join(path,'output')
    if not os.path.exists(dest_folder):
        os.mkdir('output')
        
    lst_excel = sorted([x for x in  os.listdir() if x.endswith('.xlsx')])
    for ExcelName in lst_excel:
        wb = openpyxl.load_workbook(ExcelName)
        lst_sheets = wb.sheetnames
        for sheetname  in lst_sheets:
            filename = '{}_{}.txt'.format(os.path.splitext(ExcelName)[0],sheetname)
            df  = pd.read_excel(ExcelName,sheet_name = sheetname,engine='openpyxl')
            lst = df.values.tolist()
            with open(os.path.join(dest_folder,filename), 'w') as f:
                for i in lst:
                    i = [str(x) for x in i]
                    f.write(' '.join(i) + '\n')
            print('File {} created successfully'.format(filename))
    print('Completed')
    

if __name__ == "__main__":
    path = input('Enter the Path of ExcelFiles:')
    convert_excel_txt2(path)
    